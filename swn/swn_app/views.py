from django.shortcuts import render, redirect, HttpResponse
from random import randint, choice
from .models import Sector, World
from reg_app.models import User
from matplotlib import pyplot as plt
from matplotlib import use
from matplotlib import patches as mpatches
import openpyxl
from openpyxl.utils import get_column_letter
import os

# Create your views here.

# external functions for procedural generation

def plotSector(sectorid):
    current_sector = Sector.objects.get(id=sectorid)
    worlds = World.objects.filter(sector=current_sector)
    data = {
        'all_x' : [],
        'all_y' : [],
        'all_atmos' : [],
        'all_pop' : []
    }
    use('Agg')
    
    for world in worlds:
        # coords added to data
        data['all_x'].append(world.x_coord)
        data['all_y'].append(world.y_coord)

        # atmosphere colors added to data
        if world.atmosphere == "Corrosive":
            data['all_atmos'].append('#FF7E00')
        elif world.atmosphere == "Inert":
            data['all_atmos'].append('#FFEF00')
        elif world.atmosphere == "Airless or Thin":
            data['all_atmos'].append('#000000')
        elif world.atmosphere == "Breathable":
            data['all_atmos'].append('#00BFFF')
        elif world.atmosphere == "Thick":
            data['all_atmos'].append('#A9A9A9')
        elif world.atmosphere == "Invasive":
            data['all_atmos'].append('#B0BF1A')
        elif world.atmosphere == "Corrosive and Invasive":
            data['all_atmos'].append('#D3212D')

        # population sizes added to data
        if world.population == "Failed Colony":
            data['all_pop'].append(15)
        elif world.population == "Outpost":
            data['all_pop'].append(50)
        elif world.population == "Fewer than a million":
            data['all_pop'].append(100)
        elif world.population == "Several million":
            data['all_pop'].append(150)
        elif world.population == "Hundreds of millions":
            data['all_pop'].append(200)
        elif world.population == "Billions":
            data['all_pop'].append(250)
        elif world.population == "Alien inhabitants":
            data['all_pop'].append(300)

        # create and plot figure
        plt.text(world.x_coord -0.5, world.y_coord +0.5, world.name, fontsize=6)
    plt.scatter('all_x', 'all_y', c='all_atmos', s='all_pop', marker='o', data=data)
    plt.xlabel('X Coordinates')
    plt.ylabel('Y Coordinates')
    plt.yticks([10,9,8,7,6,5,4,3,2,1])
    plt.gca().invert_yaxis()
    plt.title(f"{current_sector.name}", fontsize=20)
    plt.grid(True)

    # legend - BUG - why does this not show up as a circle?
    # REFACTOR this. There has to be a better way than custom proxy artists.
    # Also we need to show only those in the diagram
    corrosive_patch = mpatches.Circle((0.5, 0.5), 0.25, facecolor='#FF7E00', label="Corrosive")
    inert_patch = mpatches.Circle((0.5, 0.5), 0.25, facecolor='#FFEF00', label="Inert")
    airless_patch = mpatches.Circle((0.5, 0.5), 0.25, facecolor='#000000', label="Airless or Thin")
    breathable_patch = mpatches.Circle((0.5, 0.5), 0.25, facecolor='#00BFFF', label="Breathable")
    thick_patch = mpatches.Circle((0.5, 0.5), 0.25, facecolor='#A9A9A9', label="Thick")
    invasive_patch = mpatches.Circle((0.5, 0.5), 0.25, facecolor='#B0BF1A', label="Invasive")
    corrinvasive_patch = mpatches.Circle((0.5, 0.5), 0.25, facecolor='#D3212D', label="Corrosive and Invasive")
    plt.legend(
        handles=[
            corrosive_patch, 
            inert_patch, 
            airless_patch, 
            breathable_patch, 
            thick_patch, 
            invasive_patch, 
            corrinvasive_patch
        ], 
        bbox_to_anchor=(1,1), loc='upper left', scatterpoints=1)

    # file path and save
    filepath = f"media/{current_sector.id}.png"
    plt.savefig(filepath, bbox_inches='tight', ncol=2)
    plt.close()
    return

def tableLookup(roll):
    index = 0
    if (roll == 2) : index = 0
    elif (roll == 3) : index = 1
    elif (roll == 4 or roll == 5) : index = 2
    elif (roll >= 6 and roll <= 8) : index = 3
    elif (roll == 9 or roll == 10) : index = 4
    elif (roll == 11) : index = 5
    elif (roll == 12) : index = 6
    return index

def atmosLookup(roll):
    index = 0
    if (roll == 2) : index = 0
    elif (roll == 3) : index = 1
    elif (roll == 4) : index = 2
    elif (roll >= 5 and roll <= 9) : index = 3
    elif (roll == 10) : index = 4
    elif (roll == 11) : index = 5
    elif (roll == 12) : index = 6
    return index

def resolve_collision(x_coord, y_coord, sectorid): # REFACTOR this, it's janky AF, time efficiency is like O(n**2)
    worlds = World.objects.filter(sector=sectorid)
    for world in worlds:
        if x_coord == world.x_coord and y_coord == world.y_coord:
            #find nearest edge and move towards it
            check_x = (4.5 - x_coord) # determine distance from centerline
            check_y = (5.5 - y_coord) 
            print(f"Collision detected at {x_coord},{y_coord}. World {worlds.last().id + 1}.") 
            print(f"check_x = {check_x}. check_y = {check_y}.")
            # if x or y is already at an edge, move the other coordinate. Re-roll if collision in corner
            if abs(check_x) == 3.5 and abs(check_y) == 4.5: # BUG: this doesn't return properly
                x_coord = randint(1,8)
                y_coord = randint(1,10)
                print(f"World is cornered and cant be moved closer to any edge, re-rolled to {x_coord}, {y_coord}")
                return resolve_collision(x_coord, y_coord, sectorid)
            elif abs(check_x) == 3.5:
                print("x already at edge, move y instead")
                if check_y < 0:
                    y_coord += 1
                    print(f"Y coord increased to {y_coord}")
                    return resolve_collision(x_coord, y_coord, sectorid)
                else:
                    y_coord -= 1
                    print(f"Y coord decreased to {y_coord}")
                    return resolve_collision(x_coord, y_coord, sectorid)
            elif abs(check_y) == 4.5:
                print("y already at edge, move x instead")
                if check_x < 0:
                    x_coord += 1
                    print(f"X coord increased to {x_coord}")
                    return resolve_collision(x_coord, y_coord, sectorid)
                else:
                    x_coord -= 1
                    print(f"X coord decreased to {x_coord}")
                    return resolve_collision(x_coord, y_coord, sectorid)
            # we aren't on an edge, move towards nearest edge
            elif check_x < 0 and abs(check_x) >= abs(check_y):
                x_coord += 1
                print(f"X coord increased to {x_coord}")
                return resolve_collision(x_coord, y_coord, sectorid)
            elif check_x > 0 and abs(check_x) >= abs(check_y):
                x_coord -= 1
                print(f"X coord decreased to {x_coord}")
                return resolve_collision(x_coord, y_coord, sectorid)
            elif check_y < 0 and abs(check_y) > abs(check_x):
                y_coord += 1
                print(f"Y coord increased to {y_coord}")
                return resolve_collision(x_coord, y_coord, sectorid)
            elif check_y > 0 and abs(check_y) > abs(check_x):
                y_coord -= 1
                print(f"Y coord decreased to {y_coord}")
                return resolve_collision(x_coord, y_coord, sectorid)
            # special handling for any case that wasn't handled
            else:
                x_coord = randint(1,8)
                y_coord = randint(1,10)
                print("Something went wrong. Rerolling X and Y")
                return resolve_collision(x_coord, y_coord, sectorid)
    return (x_coord, y_coord)
            

def generatesector(request): # main path function for procedural generation

    # system traits dictionary
    world_traits = {
        'atmosList' : ["Corrosive", "Inert", "Airless or Thin", "Breathable", "Thick", "Invasive", "Corrosive and Invasive"],
        'tempList' : ["Frozen", "Cold", "Variable Cold", "Temperate", "Variable Warm", "Warm", "Burning"],
        'bioList' : ["Remnant", "Microbial", "No Native Biosphere", "Human Miscible", "Immiscible", "Hybrid", "Engineered"],
        'popList' : ["Failed Colony", "Outpost", "Fewer than a million", "Several million", "Hundreds of millions", "Billions", "Alien inhabitants"],
        'gravList' : ["Micro or Zero-G", "Very Low G", "Low G", "Normal", "High G", "Very High G", "Crushing Gravity"],
        'techList' : ["TL0 - Neolithic", "TL1 - Medieval", "TL2 - Early industrial", "TL3 - present-day Earth", "TL4 - Postech", "TL4+ - Postech with specialities", "TL5 - Pretech"],
    }

    # system tags list
    world_tags = [
        'Abandoned Colony', 'Flying Cities', 'Misandry/Misogyny', 'Rigid Culture',
        'Alien Ruins', 'Forbidden Tech', 'Night World', 'Rising Hegemon', 'Altered Humanity', 
        'Former Warriors', 'Nomads', 'Ritual Combat', 'Anarchists', 'Freak Geology', 
        'Oceanic World', 'Robots', 'Anthropomorphs', 'Freak Weather', 'Out of Contact',
        'Seagoing Cities', 'Area 51', 'Friendly Foe', 'Outpost World', 'Sealed Menace', 'Badlands World',
        'Gold Rush', 'Perimeter Agency', 'Secret Masters', 'Battleground', 'Great Work',
        'Pilgrimage Site', 'Sectarians', 'Beastmasters', 'Hatred', 'Pleasure World', 
        'Seismic Instability', 'Bubble Cities', 'Heavy Industry', 'Police State',
        'Shackled World', 'Cheap Life', 'Heavy Mining', 'Post-Scarcity','Societal Despair',
        'Civil War', 'Hivemind', 'Preceptor Archive', 'Sole Supplier', 'Cold War',
        'Holy War', 'Pretech Cultists', 'Taboo Treasure', 'Colonized Population', 
        'Hostile Biosphere', 'Primitive Aliens', 'Terraform Failure', 'Cultural Power',
        'Hostile Space','Prison Planet', 'Theocracy', 'Cybercommunists', 'Immortals',
        'Psionics Academy', 'Tomb World','Cyborgs', 'Local Specialty', 'Psionics Fear',
        'Trade Hub','Cyclical Doom', 'Local Tech', 'Psionics Worship', 'Tyranny',
        'Desert World', 'Major Spaceyard', 'Quarantined World', 'Unbraked AI',
        'Doomed World', 'Mandarinate','Radioactive World', 'Urbanized Surface','Dying Race',
        'Mandate Base', 'Refugees', 'Utopia', 'Eugenic Cult', 'Maneaters', 'Regional Hegemon', 
        'Warlords','Exchange Consulate', 'Megacorps','Restrictive Laws', 'Xenophiles',
        'Fallen Hegemon', 'Mercenaries','Revanchists','Xenophobes', 'Feral World',
        'Minimal Contact', 'Revolutionaries','Zombies',
    ]

    # create sector with default name
    logged_user = User.objects.get(id=request.session['userid'])
    sectors = Sector.objects.all()
    if len(sectors) == 0:
        sector_name = "Sector 1"
    else:
        sector_name = f"Sector {Sector.objects.last().id + 1}"

    Sector.objects.create(name = sector_name, owned_by = logged_user, public = False)
    current_sector = Sector.objects.last()

    # procudural generation
    num_worlds = randint(21,30)
    for i in range(num_worlds):
        
        # coordinates - random generate, check for collision and adjust
        x_coord = randint(1,8)
        y_coord = randint(1,10)
        resolution = resolve_collision(x_coord, y_coord, current_sector.id)
        x_coord, y_coord = resolution[0], resolution[1]

        # tags and traits
        tag1 = choice(world_tags)
        tag2 = choice(world_tags)
        atmosphere = world_traits['atmosList'][atmosLookup(randint(1,6) + randint(1,6))]
        gravity = world_traits['gravList'][atmosLookup(randint(1,6) + randint(1,6))]
        temperature = world_traits['tempList'][tableLookup(randint(1,6) + randint(1,6))]
        biosphere = world_traits['bioList'][tableLookup(randint(1,6) + randint(1,6))]
        population = world_traits['popList'][tableLookup(randint(1,6) + randint(1,6))]
        techlevel = world_traits['techList'][tableLookup(randint(1,6) + randint(1,6))]
        worlds= World.objects.last()
        if not worlds:
            name = "World 1"
        else:
            name = f"World {World.objects.last().id +1}"

        # create world
        World.objects.create(
            x_coord = x_coord,
            y_coord = y_coord,
            tag1 = tag1,
            tag2 = tag2,
            atmosphere = atmosphere,
            gravity = gravity,
            temperature = temperature,
            biosphere = biosphere,
            population = population,
            techlevel = techlevel,
            sector = current_sector,
            name = name
        )

    # create map plot and store in map ImageField
    plotSector(current_sector.id)
    current_sector.map = f"{current_sector.id}.png"
    current_sector.save()

    return redirect('/swn')


def sectorlist(request): #render list of sectors user can see (own or public)
    try:
        user = User.objects.get(id=request.session['userid'])
    except:
        return redirect('/')
    context = {
        'sectors' : Sector.objects.filter(owned_by=user).exclude(public=True),
        'publics' : Sector.objects.filter(public=True)
    }
    return render(request, "sectorlist.html", context)

def sectordetail(request, sectorid): #sector detail with map and table of systems
    try:
        user = User.objects.get(id=request.session['userid'])
    except:
        return redirect('/')
    current_sector = Sector.objects.get(id=sectorid)
    if current_sector.owned_by != user and current_sector.public == False:
        return redirect('/swn')
    elif current_sector.owned_by != user and current_sector.public == True:
        owned = False
    else:
        owned = True
    context = {
        'current_sector' : Sector.objects.get(id=sectorid),
        'worlds' : World.objects.filter(sector=sectorid),
        'owned' : owned,
    }
    return render(request, 'sectordetail.html', context)

def editsector(request, sectorid):
    user = User.objects.get(id=request.session['userid'])
    current_sector = Sector.objects.get(id=sectorid)
    if current_sector.owned_by != user:
        return redirect('/swn')
    else:
        context = {
            'current_sector' : Sector.objects.get(id=sectorid),
            'worlds' : World.objects.filter(sector=sectorid),
        }
        return render(request, 'editsector.html', context)

def worlddetail(request, worldid):
    user = User.objects.get(id=request.session['userid'])
    current_sector = World.objects.get(id=worldid).sector
    if current_sector.owned_by != user and current_sector.public == False:
        return redirect('/swn')
    elif current_sector.owned_by != user and current_sector.public == True:
        owned = False
    else:
        owned = True
    context = {
        'world' : World.objects.get(id=worldid),
        'owned' : owned
    }
    return render(request, 'worlddetail.html', context)

def editworld(request, worldid):
    user = User.objects.get(id=request.session['userid'])
    current_sector = World.objects.get(id=worldid).sector
    if current_sector.owned_by != user:
        return redirect('/swn')
    else:
        context = {
            'world' : World.objects.get(id=worldid)
        }
        return render(request, 'editworld.html', context)

def updatesector(request, sectorid):
    returnpath = f"/swn/sectordetail/{sectorid}"
    current_sector = Sector.objects.get(id=sectorid)
    current_sector.name = request.POST['sector_name']
    current_sector.desc = request.POST['sector_desc']
    current_sector.public = request.POST['sector_public']
    current_sector.save()
    plotSector(current_sector.id)
    return redirect(returnpath)

def updateworld(request, worldid):
    world = World.objects.get(id=worldid)
    returnpath = f"/swn/worlddetail/{world.id}"
    world.name = request.POST['world_name']
    world.x_coord = request.POST['x_coord']
    world.y_coord = request.POST['y_coord']
    world.atmosphere = request.POST['atmosphere']
    world.gravity = request.POST['gravity']
    world.temperature = request.POST['temperature']
    world.biosphere = request.POST['biosphere']
    world.population = request.POST['population']
    world.techlevel = request.POST['techlevel']
    world.tag1 = request.POST['tag1']
    world.tag2 = request.POST['tag2']
    world.desc = request.POST['desc']
    world.save()
    plotSector(world.sector.id)
    return redirect(returnpath)

def exportxlsx(request, sectorid):
    current_sector = Sector.objects.get(id=sectorid)
    worlds = World.objects.filter(sector=current_sector)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = f"{current_sector.name}"
    columns = [
        'ID',
        'Name',
        'X',
        'Y',
        'Atmosphere',
        'Gravity',
        'Temperature',
        'Biosphere',
        'Population',
        'Tech Level',
        'Tag1',
        'Tag2',
        'Desc'
    ]
    
    # set column headers
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # add map
    img_location = f"media/{current_sector.id}.png"
    img = openpyxl.drawing.image.Image(img_location)
    img.anchor='P2'
    worksheet.add_image(img)

    # set row data
    for world in worlds:
        row_num += 1
        row = [
            world.id,
            world.name,
            world.x_coord,
            world.y_coord,
            world.atmosphere,
            world.gravity,
            world.temperature,
            world.biosphere,
            world.population,
            world.techlevel,
            world.tag1,
            world.tag2,
            world.desc
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    row_num += 2
    cell = worksheet.cell(row=row_num, column=1)
    cell.value = f"Owned by: {current_sector.owned_by.first_name} {current_sector.owned_by.last_name}"
    row_num += 1
    cell = worksheet.cell(row=row_num, column=1)
    cell.value = current_sector.desc

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = f'attachment; filename="{current_sector.name}.xlsx'
    workbook.save(response)
    return response


def confirmsectordelete(request, sectorid):
    user = User.objects.get(id=request.session['userid'])
    current_sector = Sector.objects.get(id=sectorid)
    if current_sector.owned_by != user:
        return redirect('/swn')
    else:
        context = {
            'current_sector' : Sector.objects.get(id=sectorid),
            'worlds' : World.objects.filter(sector=sectorid),
        }
        return render(request, 'sectordelconf.html', context)

def confirmworlddelete(request, worldid):
    user = User.objects.get(id=request.session['userid'])
    current_sector = World.objects.get(id=worldid).sector
    if current_sector.owned_by != user:
        return redirect('/swn')
    else:
        context = {
            'world' : World.objects.get(id=worldid)
        }
        return render(request, 'worlddelconf.html', context)

def deletesector(request, sectorid):
    user = User.objects.get(id=request.session['userid'])
    current_sector = Sector.objects.get(id=sectorid)
    if current_sector.owned_by != user:
        return redirect('/swn')
    else:
        Sector.objects.get(id=sectorid).delete()
        path_to_png = f"media/{sectorid}.png"
        if os.access(path_to_png, os.F_OK):
            # This may raise an exception on a Windows server, file may be locked and in use.
            os.remove(path_to_png)
        return redirect('/swn')

def deleteworld(request, worldid):
    user = User.objects.get(id=request.session['userid'])
    current_sector = World.objects.get(id=worldid).sector
    if current_sector.owned_by != user:
        return redirect('/swn')
    else:
        world = World.objects.get(id=worldid)
        sectorid = world.sector.id
        returnpath = f"/swn/sectordetail/{sectorid}"
        World.objects.get(id=worldid).delete()
        plotSector(world.sector.id)
        return redirect(returnpath)






